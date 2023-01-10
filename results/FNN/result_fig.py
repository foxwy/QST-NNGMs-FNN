# -*- coding: utf-8 -*-
import os
os.environ['KMP_DUPLICATE_LIB_OK'] = 'TRUE'

import numpy as np
import torch
import matplotlib.pyplot as plt
import argparse
import scipy.io as scio
import openpyxl

import sys
sys.path.append(".")

plt.style.use(['science', 'no-latex'])
plt.rcParams["font.family"] = 'Arial'

font_size = 34  # 34, 28, 38
font = {'size': font_size, 'weight': 'normal'}


def Plt_set(ax, xlabel, ylabel, savepath, log_flag=0, loc=4, ncol=1, f_size=18):
    ax.tick_params(labelsize=font_size, length=5, width=5)
    ax.xaxis.set_tick_params(pad=10)
    ax.spines['bottom'].set_linewidth(5)
    ax.spines['left'].set_linewidth(5)
    ax.spines['right'].set_linewidth(5)
    ax.spines['top'].set_linewidth(5)
    font2 = {'size': f_size, 'weight': 'normal'}
    ax.legend(prop=font2, loc=loc, frameon=True, ncol=ncol)
    ax.set_xlabel(xlabel, font)  # fontweight='bold'
    ax.set_ylabel(ylabel, font)

    if log_flag == 1:
        ax.set_xscale('log')
    if log_flag == 2:
        ax.set_yscale('log')
    if log_flag == 3:
        ax.set_xscale('log')
        ax.set_yscale('log')

    if 'Time' in ylabel:
        ax.set_yticks([0.01, 0.1, 1, 10, 100])
    '''
    if 'Time' in xlabel:
        ax.set_xticks([0.1, 1, 10, 100])'''

    if 'samples' in xlabel:
        ax.set_xticks([5, 6, 7, 8, 9, 11])
        ax.set_xticklabels(['$10^5$', '$10^6$', '$10^7$', '$10^8$', '$10^9$', '$\\infty$'])
        #ax.set_xticks([7, 9, 11])
        #ax.set_xticklabels(['$10^7$', '$10^9$', '$\\infty$'])

    
    if 'samples' in ylabel:
        plt.ticklabel_format(axis="y", style="sci", scilimits=(0, 0))
        tx = ax.yaxis.get_offset_text()
        tx.set_fontsize(30)

    plt.savefig(savepath + '.pdf', dpi=600, bbox_inches='tight', pad_inches=0.05)


def Plt_set2(ax, ax2, xlabel, ylabel1, ylabel2, savepath, log_flag=0, loc=4, ncol=1, f_size=23):
    ax.tick_params(labelsize=font_size, length=5, width=5)
    ax2.tick_params(labelsize=font_size, length=5, width=5)
    ax.xaxis.set_tick_params(pad=10)
    ax.spines['bottom'].set_linewidth(5)
    ax.spines['left'].set_linewidth(5)
    ax.spines['right'].set_linewidth(5)
    ax.spines['top'].set_linewidth(5)
    font2 = {'size': 18, 'weight': 'normal'}
    ax.legend(bbox_to_anchor=(-0.02, 1.01), prop=font2, loc=3, frameon=False, ncol=1)
    ax2.legend(bbox_to_anchor=(0.52, 1.01), prop=font2, loc=3, frameon=False, ncol=1)
    #ax.legend(bbox_to_anchor=(1.22, 1.02), prop=font2, loc=2, frameon=False, ncol=1)
    #ax2.legend(bbox_to_anchor=(1.22, -0.02), prop=font2, loc=3, frameon=False, ncol=1)
    #ax.legend(prop=font2, loc=3, frameon=False, ncol=1)
    #ax2.legend(prop=font2, loc=4, frameon=False, ncol=1)
    ax.set_xlabel(xlabel, font)  # fontweight='bold'
    ax.set_ylabel(ylabel1, font)
    ax2.set_ylabel(ylabel2, font)

    if log_flag == 1:
        ax.set_xscale('log')
    if log_flag == 2:
        ax.set_yscale('log')
        ax2.set_yscale('log')
    if log_flag == 3:
        ax.set_xscale('log')
        ax.set_yscale('log')

    if 'Time' in ylabel1:
        ax.set_yticks([0.001, 0.01, 0.1, 1, 10, 100, 1000])

    if 'Iteration' in ylabel2:
        ax2.set_yticks([1, 10, 100, 1000, 10000])
        #ax2.set_yticks([700, 800, 900, 1000])

    if 'samples' in xlabel:
        ax.set_xticks([5, 6, 7, 8, 9, 11])
        ax.set_xticklabels(['$10^5$', '$10^6$', '$10^7$', '$10^8$', '$10^9$', '$\\infty$'])
        #ax.set_xticks([9, 11])
        #ax.set_xticklabels(['$10^9$', '$\\infty$'])

    plt.savefig(savepath + '.pdf', dpi=600, bbox_inches='tight', pad_inches=0.05)


def Get_r(V):
    V_avg = np.mean(V, 0)
    V_std = np.std(V, 0) / 2
    r1 = list(map(lambda x: x[0] - x[1], zip(V_avg, V_std)))
    r2 = list(map(lambda x: x[0] + x[1], zip(V_avg, V_std)))

    return r1, r2


def Pure_ex_fig(colors):
    wb = openpyxl.load_workbook('result/Pure_state_sample_qubit_099.xlsx')
    ws = wb['Sheet1']

    N_q = []
    GHZi = []
    GHZi_RNN = []
    Product = []
    Product_RNN = []
    for N, S1, S1_r, S2, S2_r in zip(ws['A'][2:9], ws['B'][2:9], ws['B'][10:17], ws['B'][18:25], ws['B'][26:33]):
        N_q.append(N.value)
        GHZi.append(S1.value)
        GHZi_RNN.append(S1_r.value)
        Product.append(S2.value)
        Product_RNN.append(S2_r.value)

    fig, ax = plt.subplots(1, 1, figsize=(8, 7))
    ax.plot(N_q, GHZi, label='$\\rm GHZi$', linewidth=5, color=colors[9])
    ax.plot(N_q, GHZi_RNN, label='$\\rm GHZi+RNN$', linewidth=5, color=colors[9], marker='*', markersize=15)
    ax.plot(N_q, Product, label='$\\rm Product$', linewidth=5, color=colors[0])
    ax.plot(N_q, Product_RNN, label='$\\rm Product+RNN$', linewidth=5, color=colors[0], marker='*', markersize=15)

    ax.set_yticks([0, 5000, 10000, 15000, 20000, 25000, 30000])
    ax.set_xticks([2, 3, 4, 5, 6, 7, 8])
    Plt_set(ax, 'Number of qubits', 'Number of samples', 'fig/pure_state_sample_qubit', loc=2)
    #plt.show()


#-----ex: 1 (Random State Convergence Experiments of NN-QST for Different Loss Functions)-----
def Loss_ex_fig(na_state, r_path, Algo, colors):
    N_Samples = np.array([7, 9, 11])
    labels = ['NLL', 'SH', 'PC', 'MSE', 'NLL_PC'] 
    m_methods = 'chol_h'

    #-----sample, Fq-----
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    
    for i, loss in enumerate(labels):
        Fq_m = []
        for N_s in N_Samples:
            savePath = r_path + Algo + '_' + m_methods + '_' + str(N_s) + '_' + loss
            print(savePath)

            results = np.load(savePath + '.npy', allow_pickle=True).item()
            Fq = []
            for p in results:
                Fq_t = np.minimum(torch.tensor(results[p]['Fq']).numpy(), 1)
                Fq.append(Fq_t[-1])

            Fq_m.append(Fq)

        Fq_m = np.array(Fq_m).T

        Fq_avg = np.mean(Fq_m, 0)
        r1, r2 = Get_r(Fq_m)
        r1 = np.maximum(r1, 0)
        r2 = np.minimum(r2, 1)

        ax.plot(N_Samples, Fq_avg, linewidth=5, label=labels[i], color=colors[i])
        ax.fill_between(N_Samples, r1, r2, alpha=0.1, color=colors[i])

    ax.plot([N_Samples[0], N_Samples[-1]], [1, 1], 'k--', linewidth=5)
    ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1])
    #ax.set_yticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])
    Plt_set(ax, "Number of samples", "Quantum fidelity", 'fig/'+na_state+'/loss/'+Algo+'_'+na_state+'_sample_fq', 0, ncol=2)


    #-----purity, Fq-----
    for N_s in [7, 9, 11]:
        fig, ax = plt.subplots(1, 1, figsize=(9, 7))
        #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
        for i, loss in enumerate(labels):
            savePath = r_path + Algo + '_' + m_methods + '_' + str(N_s) + '_' + loss
            print(savePath)

            results = np.load(savePath + '.npy', allow_pickle=True).item()

            Fq_NN = []
            P_NN = []
            for p in results:
                P_NN.append(float(p))
            P_NN.sort()
            for p in P_NN:
                p = str(p)
                Fq_t = np.minimum(torch.tensor(results[p]['Fq']).numpy(), 1)
                Fq_NN.append(Fq_t[-1])

            ax.plot(np.array(P_NN)**2*(1-1/2**10)+1/2**10, Fq_NN, linewidth=5, label=labels[i], color=colors[i])

        ax.set_xticks([0, 0.2, 0.4, 0.6, 0.8, 1])
        #ax.set_xticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])
        ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1])
        #ax.set_yticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])

        Plt_set(ax, "Purity", "Quantum fidelity", 'fig/'+na_state+'/loss/'+Algo+'_'+na_state+'_'+str(N_s)+'_purity_fq', ncol=2)

#-----ex: 2 (Random State Convergence Experiments of NN-QST for Different Mapping Methods)-----
def Map_ex_fig(na_state, r_path, Algo, colors):
    N_Samples = np.array([5, 6, 7, 8, 9, 11])
    m_methods = ['chol', 'chol_h', 'proj_F', 'proj_S', 'proj_A_0.5', 'proj_A_1', 'proj_A_1.5', 'proj_A_3', 'proj_A_4']
    labels = ['$\\rm Chol_\Delta$', '$\\rm Chol_H$', '$\\rm \mathcal{F}[\cdot]$', '$\\rm \mathcal{S}[\cdot]$', \
              '$\\rm \mathcal{A}[\cdot]_{0.5}$', '$\\rm \mathcal{A}[\cdot]_1$', '$\\rm \mathcal{A}[\cdot]_{1.5}$', \
              '$\\rm \mathcal{A}[\cdot]_3$', '$\\rm \mathcal{A}[\cdot]_4$']

    #-----sample, Fq-----
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    for i in range(len(m_methods)):
        print(Algo, m_methods[i])

        Fq_m = []
        for N_s in N_Samples:
            savePath = r_path + Algo + '_' + m_methods[i] + '_' + str(N_s)

            results = np.load(savePath + '.npy', allow_pickle=True).item()
            Fq = []
            for p in results:
                Fq_t = np.minimum(torch.tensor(results[p]['Fq']).numpy(), 1)
                Fq.append(Fq_t[-1])

            Fq_m.append(Fq)


        Fq_m = np.array(Fq_m).T

        Fq_avg = np.mean(Fq_m, 0)
        r1, r2 = Get_r(Fq_m)
        r1 = np.maximum(r1, 0)
        r2 = np.minimum(r2, 1)

        ax.plot(N_Samples, Fq_avg, linewidth=5, label=labels[i], color=colors[i])
        ax.fill_between(N_Samples, r1, r2, alpha=0.1, color=colors[i])

    ax.plot([N_Samples[0], N_Samples[-1]], [1, 1], 'k--', linewidth=5)
    ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1])
    #ax.set_yticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])
    Plt_set(ax, "Number of samples", "Quantum fidelity", 'fig/'+na_state+'/map/'+Algo+'_'+na_state+'_sample_fq', 0, ncol=2)


    #-----purity, Fq-----
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    for i in range(len(m_methods)):
        print(Algo, m_methods[i])
        savePath = r_path + Algo + '_' + m_methods[i] + '_' + str(11)

        results = np.load(savePath + '.npy', allow_pickle=True).item()

        Fq_NN = []
        P_NN = []
        for p in results:
            P_NN.append(float(p))
        P_NN.sort()
        for p in P_NN:
            p = str(p)
            Fq_t = np.minimum(torch.tensor(results[p]['Fq']).numpy(), 1)
            Fq_NN.append(Fq_t[-1])

        ax.plot(np.array(P_NN)**2*(1-1/2**10)+1/2**10, Fq_NN, linewidth=5, label=labels[i], color=colors[i])

    ax.set_xticks([0, 0.2, 0.4, 0.6, 0.8, 1])
    #ax.set_xticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])
    ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1])
    #ax.set_yticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])

    Plt_set(ax, "Purity", "Quantum fidelity", 'fig/'+na_state+'/map/'+Algo+'_'+na_state+'_purity_fq', ncol=2)

#-----ex: 3 (Random State Convergence Experiments of QST algorithms for Different samples)-----
def Conv_sample_ex_fig(na_state, r_path, Algo, colors):
    N_Samples = np.array([5, 6, 7, 8, 9, 11])
    m_methods = ['NN', 'MLE', 'iMLE', 'APG', 'LRE', 'APG', 'LRE']
    labels = ['NN', 'SNN', 'iMLE', 'CG_APG', 'LRE', 'CG_APG_$\\rm \mathcal{A}[\cdot]_3$', 'LRE_$\\rm \mathcal{A}[\cdot]_1$']
    colors_o = colors

    #-----sample, Fq-----
    '''
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    for i in range(len(m_methods[:5])):
        print(m_methods[i])

        Fq_m = []
        for N_s in N_Samples:
            savePath = r_path + str(N_s)

            if N_s == 8 and m_methods[i] == 'iMLE':
                savePath = r_path + '8_1'
            results = np.load(savePath + '.npy', allow_pickle=True).item()
            Fq = []
            for p in results:
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)
                Fq.append(Fq_t[-1])

            Fq_m.append(Fq)


        Fq_m = np.array(Fq_m).T

        Fq_avg = np.mean(Fq_m, 0)
        r1, r2 = Get_r(Fq_m)
        r1 = np.maximum(r1, 0)
        r2 = np.minimum(r2, 1)

        ax.plot(N_Samples, Fq_avg, linewidth=5, label=labels[i], color=colors[i])
        ax.fill_between(N_Samples, r1, r2, alpha=0.1, color=colors[i])

    m_methods = m_methods[5:]
    labels = labels[5:]
    colors = colors[3:]
    for i in range(len(m_methods)):
        print(m_methods[i])

        Fq_m = []
        for N_s in N_Samples:
            savePath = r_path + str(N_s)

            results = np.load(savePath + '_2.npy', allow_pickle=True).item()
            Fq = []
            for p in results:
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)
                Fq.append(Fq_t[-1])

            Fq_m.append(Fq)


        Fq_m = np.array(Fq_m).T

        Fq_avg = np.mean(Fq_m, 0)
        r1, r2 = Get_r(Fq_m)
        r1 = np.maximum(r1, 0)
        r2 = np.minimum(r2, 1)

        ax.plot(N_Samples, Fq_avg, '--', linewidth=5, label=labels[i], color=colors[i], marker='o', markersize=10)
        ax.fill_between(N_Samples, r1, r2, alpha=0.1, color=colors[i])

    ax.plot([N_Samples[0], N_Samples[-1]], [1, 1], 'k--', linewidth=5)
    ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1])
    #ax.set_yticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])
    Plt_set(ax, "Number of samples", "Quantum fidelity", 'fig/'+na_state+'/sample/'+na_state+'_sample_fq', 0, ncol=2)


    #-----purity, Fq-----
    for N_s in N_Samples:
        m_methods = ['NN', 'MLE', 'iMLE', 'APG', 'LRE', 'APG', 'LRE']
        labels = ['NN', 'SNN', 'iMLE', 'CG_APG', 'LRE', 'CG_APG_$\\rm \mathcal{A}[\cdot]_3$', 'LRE_$\\rm \mathcal{A}[\cdot]_1$']
        colors = colors_o
        print(N_s)
        fig, ax = plt.subplots(1, 1, figsize=(9, 7))
        #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
        for i in range(len(m_methods[:5])):
            savePath = r_path + str(N_s)

            if N_s == 8 and m_methods[i] == 'iMLE':
                savePath = r_path + '8_1'

            results = np.load(savePath + '.npy', allow_pickle=True).item()

            Fq_NN = []
            P_NN = []
            for p in results:
                P_NN.append(float(p))
            P_NN.sort()
            for p in P_NN:
                p = str(p)
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)
                Fq_NN.append(Fq_t[-1])

            ax.plot(np.array(P_NN)**2*(1-1/2**10)+1/2**10, Fq_NN, linewidth=5, label=labels[i], color=colors[i])

        m_methods = m_methods[5:]
        labels = labels[5:]
        colors = colors[3:]
        for i in range(len(m_methods)):
            savePath = r_path + str(N_s)

            results = np.load(savePath + '_2.npy', allow_pickle=True).item()

            Fq_NN = []
            P_NN = []
            for p in results:
                P_NN.append(float(p))
            P_NN.sort()
            for p in P_NN:
                p = str(p)
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)
                Fq_NN.append(Fq_t[-1])

            ax.plot(np.array(P_NN)**2*(1-1/2**10)+1/2**10, Fq_NN, '--', linewidth=5, label=labels[i], color=colors[i], marker='o', markersize=10)

        ax.set_xticks([0, 0.2, 0.4, 0.6, 0.8, 1])
        #ax.set_xticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])
        ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1])
        #ax.set_yticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])

        Plt_set(ax, "Purity", "Quantum fidelity", 'fig/'+na_state+'/sample/'+na_state+'_'+str(N_s)+'_purity_fq', ncol=2)'''


    #-----sample, time-----
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    m_methods = ['NN', 'MLE', 'iMLE', 'APG', 'LRE', 'APG', 'LRE']
    labels_t = ['Time of NN', 'Time of SNN', 'Time of iMLE', 'Time of CG_APG', 'Time of LRE', 'Time of CG_APG_$\\rm \mathcal{A}[\cdot]_3$', 'Time of LRE_$\\rm \mathcal{A}[\cdot]_1$']
    labels_e = ['Iteration of NN', 'Iteration of SNN', 'Iteration of iMLE', 'Iteration of CG_APG', 'Iteration of LRE', 'Iteration of CG_APG_$\\rm \mathcal{A}[\cdot]_3$', 'Iteration of LRE_$\\rm \mathcal{A}[\cdot]_1$']
    ax2 = ax.twinx()
    colors = colors_o
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    for i in range(len(m_methods)):
        print(m_methods[i])

        Time_m = []
        Epoch_m = []
        for N_s in N_Samples:
            savePath = r_path + str(N_s)

            if N_s == 8 and m_methods[i] == 'iMLE':
                savePath = r_path + '8_1'

            if i >= 5:
                results = np.load(savePath + '_2.npy', allow_pickle=True).item()
            else:
                results = np.load(savePath + '.npy', allow_pickle=True).item()
            Time = []
            Epoch = []
            for p in results:
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)

                for j in range(len(Fq_t)):
                    if Fq_t[j] >= 0.99:
                        Time.append(results[p][m_methods[i]]['time'][j])
                        if m_methods[i] == 'LRE':
                            Epoch.append(1)
                        else:
                            Epoch.append(results[p][m_methods[i]]['epoch'][j])
                        break

                    if j == len(Fq_t) - 1:
                        Time.append(results[p][m_methods[i]]['time'][j])
                        if m_methods[i] == 'LRE':
                            Epoch.append(1)
                        else:
                            Epoch.append(results[p][m_methods[i]]['epoch'][j])

            if m_methods[i] == 'MLE' or m_methods[i] == 'NN':
                results_t = np.load(savePath + '_NN_SNN.npy', allow_pickle=True).item()
                Time = []
                for p in results_t:
                    Fq_t = np.minimum(torch.tensor(results_t[p][m_methods[i]]['Fq']).numpy(), 1)

                    for j in range(len(Fq_t)):
                        if Fq_t[j] >= 0.99:
                            Time.append(results_t[p][m_methods[i]]['time'][j])
                            break

                        if j == len(Fq_t) - 1:
                            Time.append(results_t[p][m_methods[i]]['time'][j])


            Time_m.append(Time)
            Epoch_m.append(Epoch)


        Time_m = np.array(Time_m).T
        Time_avg = np.mean(Time_m, 0)

        Epoch_m = np.array(Epoch_m).T
        Epoch_avg = np.mean(Epoch_m, 0)
        
        '''
        r1, r2 = Get_r(Time_m)
        r1 = np.maximum(r1, 0)
        r2 = np.minimum(r2, 1)'''

        ax.plot(N_Samples, Time_avg, linewidth=5, label=labels_t[i], color=colors[i], marker='^', markersize=10)
        ax2.plot(N_Samples, Epoch_avg, '--', linewidth=5, label=labels_e[i], color=colors[i], marker='o', markersize=10)
        #ax.fill_between(N_Samples, r1, r2, alpha=0.1, color=colors[i])

    Plt_set2(ax, ax2, "Number of samples", "Time ($s$)", 'Iteration', 'fig/'+na_state+'/sample/'+na_state+'_sample_time', 3, ncol=2)

#-----ex: 4 (Random State Convergence Experiments of QST algorithms for Different samples)-----
def Conv_pretrain_ex_fig(na_state, r_path, Algo, colors):
    N_Samples = np.array([8, 9, 11])
    #m_methods = ['NN_MLE', 'MLE', 'iMLE', 'APG', 'LRE', 'APG_projA', 'LRE_projA']
    #labels = ['Pretrain_NN_SNN', 'SNN', 'iMLE', 'CG_APG', 'LRE', '$\\rm \mathcal{A}[\cdot]_3$', '$\\rm \mathcal{A}[\cdot]_1']
    m_methods = ['NN_MLE', 'MLE']
    labels = ['Pretrain_NN_SNN', 'SNN']

    #-----sample, Fq-----
    '''
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    for i in range(len(m_methods)):
        print(m_methods[i])

        Fq_m = []
        for N_s in N_Samples:
            savePath = r_path + str(N_s)

            results = np.load(savePath + '_pretrain.npy', allow_pickle=True).item()
            Fq = []
            for p in results:
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)
                Fq.append(Fq_t[-1])

            Fq_m.append(Fq)


        Fq_m = np.array(Fq_m).T

        Fq_avg = np.mean(Fq_m, 0)
        r1, r2 = Get_r(Fq_m)
        r1 = np.maximum(r1, 0)
        r2 = np.minimum(r2, 1)

        ax.plot(N_Samples, Fq_avg, linewidth=5, label=labels[i], color=colors[i])
        ax.fill_between(N_Samples, r1, r2, alpha=0.1, color=colors[i])

    ax.plot([N_Samples[0], N_Samples[-1]], [1, 1], 'k--', linewidth=5)
    ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1])
    #ax.set_yticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])
    Plt_set(ax, "Number of samples", "Quantum fidelity", 'fig/'+na_state+'/pretrain/'+na_state+'_sample_fq', 0, ncol=2)'''


    #-----purity, Time, Epoch-----
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    #labels_t = ['Time of Pretrain_NN_SNN', 'Time of SNN', 'Time of iMLE', 'Time of CG_APG', 'Time of LRE', 'Time of CG_APG_$\\rm \mathcal{A}[\cdot]_3$', 'Time of LRE_$\\rm \mathcal{A}[\cdot]_1$']
    #labels_e = ['Iteration of Pretrain_NN_SNN', 'Iteration of SNN',  'Iteration of iMLE', 'Iteration of CG_APG', 'Iteration of LRE', 'Iteration of CG_APG_$\\rm \mathcal{A}[\cdot]_3$', 'Iteration of LRE_$\\rm \mathcal{A}[\cdot]_1$']
    labels_t = ['Time of Pretrain_NN_SNN', 'Time of SNN']
    labels_e = ['Iteration of Pretrain_NN_SNN', 'Iteration of SNN']

    ax2 = ax.twinx()
    for i in range(len(m_methods)):
        savePath = r_path + str(11)

        results = np.load(savePath + '_pretrain_6.npy', allow_pickle=True).item()

        P_NN = []
        Time = []
        Epoch = []
        for p in results:
            P_NN.append(float(p))
        P_NN.sort()
        for p in P_NN:
            p = str(p)
            Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)
            
            for j in range(len(Fq_t)):
                if Fq_t[j] >= 0.99:
                    Time.append(results[p][m_methods[i]]['time'][j])
                    if 'LRE' in m_methods[i]:
                        Epoch.append(1)
                    else:
                        Epoch.append(results[p][m_methods[i]]['epoch'][j])
                    break

                if j == len(Fq_t) - 1:
                    Time.append(results[p][m_methods[i]]['time'][j])
                    if 'LRE' in m_methods[i]:
                        Epoch.append(1)
                    else:
                        Epoch.append(results[p][m_methods[i]]['epoch'][j])

        ax.plot(np.array(P_NN)**2*(1-1/2**7)+1/2**7, Time, linewidth=5, label=labels_t[i], color=colors[i], marker='^', markersize=10)
        ax2.plot(np.array(P_NN)**2*(1-1/2**7)+1/2**7, Epoch, '--', linewidth=5, label=labels_e[i], color=colors[i], marker='o', markersize=10)
        print(Epoch)

    ax.set_xticks([0, 0.2, 0.4, 0.6, 0.8, 1])
    #ax.set_xticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])

    Plt_set2(ax, ax2, "Purity", "Time ($s$)", 'Iteration', 'fig/'+na_state+'/pretrain/'+na_state+'_purity_time', 2, ncol=2)
    #plt.show()


    #-----sample, time-----
    '''
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    labels_t = ['Time of NN_SNN', 'Time of SNN', 'Time of iMLE', 'Time of CG_APG', 'Time of LRE', 'Time of CG_APG_projA', 'Time of LRE_projA']
    labels_e = ['Epoch of NN_SNN', 'Epoch of SNN',  'Epoch of iMLE', 'Epoch of CG_APG', 'Epoch of LRE', 'Epoch of CG_APG_projA', 'Epoch of LRE_projA']
    ax2 = ax.twinx()
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    for i in range(len(m_methods)):
        print(m_methods[i])

        Time_m = []
        Epoch_m = []
        for N_s in N_Samples:
            savePath = r_path + str(N_s)

            results = np.load(savePath + '_pretrain.npy', allow_pickle=True).item()
            Time = []
            Epoch = []
            for p in results:
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)

                for j in range(len(Fq_t)):
                    if Fq_t[j] >= 0.99:
                        Time.append(results[p][m_methods[i]]['time'][j])
                        if 'LRE' in m_methods[i]:
                            Epoch.append(1)
                        else:
                            Epoch.append(results[p][m_methods[i]]['epoch'][j])
                        break

                    if j == len(Fq_t) - 1:
                        Time.append(results[p][m_methods[i]]['time'][j])
                        if 'LRE' in m_methods[i]:
                            Epoch.append(1)
                        else:
                            Epoch.append(results[p][m_methods[i]]['epoch'][j])

            Time_m.append(Time)
            Epoch_m.append(Epoch)


        Time_m = np.array(Time_m).T
        Time_avg = np.mean(Time_m, 0)

        Epoch_m = np.array(Epoch_m).T
        Epoch_avg = np.mean(Epoch_m, 0)
        
        r1, r2 = Get_r(Time_m)
        r1 = np.maximum(r1, 0)
        r2 = np.minimum(r2, 1)

        ax.plot(N_Samples, Time_avg, linewidth=5, label=labels_t[i], color=colors[i], marker='^', markersize=10)
        ax2.plot(N_Samples, Epoch_avg, '--', linewidth=5, label=labels_e[i], color=colors[i], marker='o', markersize=10)
        #ax.fill_between(N_Samples, r1, r2, alpha=0.1, color=colors[i])

    Plt_set2(ax, ax2, "Number of samples", "Time ($s$)", 'Epoch', 'fig/'+na_state+'/pretrain/'+na_state+'_sample_time', 3, ncol=2)'''

#-----ex: 5 (Random State Convergence Experiments of QST algorithms for Different samples)-----
def Conv_CNN_ex_fig(na_state, r_path, Algo, colors):
    N_Samples = np.array([5, 6, 7, 8, 9, 11])
    m_methods = ['NN', 'CNN']
    labels = ['FNN', 'CNN']

    #-----sample, Fq-----
    '''
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    for i in range(len(m_methods)):
        print(m_methods[i])

        Fq_m = []
        for N_s in N_Samples:
            savePath = r_path + str(N_s)

            results = np.load(savePath + '_NN_CNN.npy', allow_pickle=True).item()
            Fq = []
            for p in results:
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)
                Fq.append(Fq_t[-1])

            Fq_m.append(Fq)


        Fq_m = np.array(Fq_m).T

        Fq_avg = np.mean(Fq_m, 0)
        r1, r2 = Get_r(Fq_m)
        r1 = np.maximum(r1, 0)
        r2 = np.minimum(r2, 1)

        ax.plot(N_Samples, Fq_avg, linewidth=5, label=labels[i], color=colors[i])
        ax.fill_between(N_Samples, r1, r2, alpha=0.1, color=colors[i])

    ax.plot([N_Samples[0], N_Samples[-1]], [1, 1], 'k--', linewidth=5)
    ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1])
    #ax.set_yticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])
    Plt_set(ax, "Number of samples", "Quantum fidelity", 'fig/'+na_state+'/sample/'+na_state+'_FNN_CNN_sample_fq', 0, ncol=2)'''

    #-----sample, time-----
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    labels_t = ['Time of FNN', 'Time of CNN']
    labels_e = ['Iteration of FNN', 'Iteration of CNN']
    ax2 = ax.twinx()
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    for i in range(len(m_methods)):
        print(m_methods[i])

        Time_m = []
        Epoch_m = []
        for N_s in N_Samples:
            savePath = r_path + str(N_s)

            results = np.load(savePath + '_NN_CNN.npy', allow_pickle=True).item()
            Time = []
            Epoch = []
            for p in results:
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)

                for j in range(len(Fq_t)):
                    if Fq_t[j] >= 0.99:
                        Time.append(results[p][m_methods[i]]['time'][j])
                        if m_methods[i] == 'LRE':
                            Epoch.append(1)
                        else:
                            Epoch.append(results[p][m_methods[i]]['epoch'][j])
                        break

                    if j == len(Fq_t) - 1:
                        Time.append(results[p][m_methods[i]]['time'][j])
                        if m_methods[i] == 'LRE':
                            Epoch.append(1)
                        else:
                            Epoch.append(results[p][m_methods[i]]['epoch'][j])

            Time_m.append(Time)
            Epoch_m.append(Epoch)


        Time_m = np.array(Time_m).T
        Time_avg = np.mean(Time_m, 0)

        Epoch_m = np.array(Epoch_m).T
        Epoch_avg = np.mean(Epoch_m, 0)
        
        ax.plot(N_Samples, Time_avg, linewidth=5, label=labels_t[i], color=colors[i], marker='^', markersize=10)
        ax2.plot(N_Samples, Epoch_avg, '--', linewidth=5, label=labels_e[i], color=colors[i], marker='o', markersize=10)

    Plt_set2(ax, ax2, "Number of samples", "Time ($s$)", 'Iteration', 'fig/'+na_state+'/sample/'+na_state+'_FNN_CNN_sample_time', 3, ncol=2)

#-----ex: 6 (Convergence Experiment of Random Mixed States for Different Qubits)-----
def Conv_qubit_ex_fig(na_state, r_path, Algo, colors):
    m_methods = ['NN', 'MLE', 'iMLE', 'APG', 'LRE', 'APG_projA', 'LRE_projA']
    labels = ['NN', 'SNN', 'iMLE', 'CG_APG', 'LRE', 'CG_APG_$\\rm \mathcal{A}[\cdot]_3$', 'LRE_$\\rm \mathcal{A}[\cdot]_1$']
    colors_o = colors

    #-----sample, time-----
    fig, ax = plt.subplots(1, 1, figsize=(9, 7))
    labels_t = ['Time of NN', 'Time of SNN', 'Time of iMLE', 'Time of CG_APG', 'Time of LRE', 'Time of CG_APG_$\\rm \mathcal{A}[\cdot]_3$', 'Time of LRE_$\\rm \mathcal{A}[\cdot]_1$']
    labels_e = ['Iteration of NN', 'Iteration of SNN', 'Iteration of iMLE', 'Iteration of CG_APG', 'Iteration of LRE', 'Iteration of CG_APG_$\\rm \mathcal{A}[\cdot]_3$', 'Iteration of LRE_$\\rm \mathcal{A}[\cdot]_1$']
    ax2 = ax.twinx()
    colors = colors_o
    #markers = ['.', '.', '*', '*', '^', '^', '^', '^']
    for i in range(len(m_methods)):
        print(m_methods[i])

        Time_m = []
        Epoch_m = []
        if m_methods[i] == 'APG_projA' or m_methods[i] == 'iMLE':
            N_qubits = np.arange(2, 10)
        else:
            N_qubits = np.arange(2, 12)

        for N_q in N_qubits:
            savePath = r_path + str(N_q)

            if 'NN' in labels[i]:
                results = np.load(savePath + 'q_11NN.npy', allow_pickle=True).item()
            else:
                results = np.load(savePath + 'q_11.npy', allow_pickle=True).item()
            Time = []
            Epoch = []
            for p in results:
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)

                for j in range(len(Fq_t)):
                    if Fq_t[j] >= 0.99:
                        Time.append(results[p][m_methods[i]]['time'][j])
                        if 'LRE' in m_methods[i]:
                            Epoch.append(1)
                        else:
                            Epoch.append(results[p][m_methods[i]]['epoch'][j])
                        break

                    if j == len(Fq_t) - 1:
                        Time.append(results[p][m_methods[i]]['time'][j])
                        if 'LRE' in m_methods[i]:
                            Epoch.append(1)
                        else:
                            Epoch.append(results[p][m_methods[i]]['epoch'][j])

            Time_m.append(Time)
            Epoch_m.append(Epoch)


        Time_m = np.array(Time_m).T
        Time_avg = np.mean(Time_m, 0)

        Epoch_m = np.array(Epoch_m).T
        Epoch_avg = np.mean(Epoch_m, 0)

        ax.plot(N_qubits, Time_avg, linewidth=5, label=labels_t[i], color=colors[i], marker='^', markersize=10)
        ax2.plot(N_qubits, Epoch_avg, '--', linewidth=5, label=labels_e[i], color=colors[i], marker='o', markersize=10)
        print(Time_avg, Epoch_avg)

    ax.set_xticks(np.arange(2, 12))
    Plt_set2(ax, ax2, "Number of qubits", "Time ($s$)", 'Iteration', 'fig/'+na_state+'/qubit/'+na_state+'_qubit_time', log_flag=2, ncol=2)
    
#-----ex: 7 (Convergence Experiment of Random Mixed States for Different Samples)-----
def Conv_depolar_ex_fig(na_state, r_path, Alpha, colors):
    N_Samples = np.array([9])
    m_methods = ['NN', 'MLE', 'iMLE', 'APG', 'LRE', 'APG', 'LRE']
    labels = ['NN', 'SNN', 'iMLE', 'CG_APG', 'LRE', 'CG_APG_$\\rm \mathcal{A}[\cdot]_3$', 'LRE_$\\rm \mathcal{A}[\cdot]_1$']
    #-----purity, Fq-----
    for N_s in N_Samples:
        print(N_s)
        fig, ax = plt.subplots(1, 1, figsize=(9, 7))
        for i in range(len(m_methods)):
            savePath = r_path + str(N_s)

            results = np.load(savePath + '_depolar.npy', allow_pickle=True).item()

            Fq_NN = []
            P_NN = []
            for p in results:
                P_NN.append(float(p))
            P_NN.sort()
            for p in P_NN:
                p = str(p)
                Fq_t = np.minimum(torch.tensor(results[p][m_methods[i]]['Fq']).numpy(), 1)
                Fq_NN.append(Fq_t[-1])

            ax.plot(np.array(P_NN), Fq_NN, linewidth=5, label=labels[i], color=colors[i])

        ax.plot(np.array(P_NN), 1 - (1 - 1 / 2**10) * np.array(P_NN), 'k--', linewidth=5)
        ax.set_xticks([0, 0.2, 0.4, 0.6, 0.8, 1])
        #ax.set_xticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])
        ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1])
        #ax.set_yticklabels(['$0.0$', '$0.2$', '$0.4$', '$0.6$', '$0.8$', '$1.0$'])

        Plt_set(ax, "$\lambda$", "Quantum fidelity", 'fig/'+na_state+'/depolar/'+na_state+'_'+str(N_s)+'_depolar_lambda_fq', loc=1, ncol=1)


if __name__ == '__main__':
    colors = ['#75bbfd', 'magenta', '#658b38', '#c79fef', '#06c2ac', 'cyan', '#ff9408', '#430541', 'blue', '#fb2943']
    na_state = 'real_random'

    Algo = 'NN'

    print('-----state:', na_state)
    r_path = 'result/' + na_state + '/'
    Conv_depolar_ex_fig(na_state, r_path, Algo, colors)